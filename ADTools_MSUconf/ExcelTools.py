import numpy as np
from numpy.linalg import norm 
####################################
import openpyxl
from openpyxl.styles import Font, Style
import os.path
####################################
import itertools

def GetString(iChar, iNum):
    iNum += ord(iChar)
    first = ord('A')
    last = ord('Z')
    str = ''
    iNum = iNum-first
    if iNum == 0: return 'A'
    while iNum!=0:
        (iNum,rem)=divmod(iNum,last-first+1)
        if str:
            str+=chr(rem-1+first)
        else:
            str+=chr(rem+first)
    return str[::-1]
        
def InitializeFirstSheet(wb,sname,ListOfMethods):
    sheet = wb.get_sheet_by_name('{}'.format(sname))
    #Main Table
    sheet.merge_cells('A1:A2')
    sheet['A1'].style = Style(font=Font(bold=True))
    sheet['A1'] = 'N'
    for i in range(len(ListOfMethods)):
        i4PlusBchar = GetString('B',i*4)
        i4PlusCchar = GetString('C',i*4)
        i4PlusDchar = GetString('D',i*4)
        i4PlusEchar = GetString('E',i*4)
        sheet.merge_cells('{}1:{}1'.format(i4PlusBchar,i4PlusEchar))
        sheet['{}1'.format(i4PlusBchar)].style = Style(font=Font(bold=True))
        sheet['{}1'.format(i4PlusBchar)] = '{}'.format(ListOfMethods[i])
        sheet['{}2'.format(i4PlusBchar)] = 'Initial Energy'
        sheet['{}2'.format(i4PlusCchar)] = 'Gradient Norm'
        sheet['{}2'.format(i4PlusDchar)] = 'Resultant Energy'
        sheet['{}2'.format(i4PlusEchar)] = 'Gradient Norm'
    
def SaveResults(wb, sname, ListOfMethods, N, Ein, nGradin, Eopt, nGradopt):
    sheet = wb.get_sheet_by_name('{}'.format(sname))
    #Main Table
    sheet['A{}'.format(N)] = N
    for i in range(len(ListOfMethods)):
        i4PlusBchar = GetString('B',i*4)
        i4PlusCchar = GetString('C',i*4)
        i4PlusDchar = GetString('D',i*4)
        i4PlusEchar = GetString('E',i*4)
        sheet['{}{}'.format(i4PlusBchar, N)] = Ein[i]
        sheet['{}{}'.format(i4PlusCchar, N)] = nGradin[i]
        sheet['{}{}'.format(i4PlusDchar, N)] = Eopt[i]
        sheet['{}{}'.format(i4PlusEchar, N)] = nGradopt[i]
        
def SavePrecisionToExcel(fname, ListOfMethods,ListOfValues, N, Ein, nGradin, Eopt, nGradopt):
    nofmethods = len(ListOfMethods)
    ###################################################
    # Write data to an excel file
    # Step 1. Initialization
    ###################################################
    FirstSheetName = fname.split('.')[0]
    
    if(os.path.isfile(fname)):     
        wb = openpyxl.load_workbook(fname)
        snames = wb.get_sheet_names()
        wb.create_sheet(index=len(snames), title='Exp{} N={}'.format(len(snames),N))
        sheet = wb.get_sheet_by_name('Exp{} N={}'.format(len(snames),N))
    else:
        wb = openpyxl.Workbook()
        #First sheet is the main one for the results
        sheet = wb.active
        sheet.title='{}'.format(fname.split('.')[0])
        InitializeFirstSheet(wb,FirstSheetName,ListOfMethods)
        #Second and others are for storing absolute and relative errors
        snames = wb.get_sheet_names()
        wb.create_sheet(index=len(snames), title='Exp1 N={}'.format(N))
        sheet = wb.get_sheet_by_name('Exp1 N={}'.format(N))
    
    SaveResults(wb, FirstSheetName, ListOfMethods, N, Ein, nGradin, Eopt, nGradopt)  
    #Table Parameters
    #Change it if it is needed 
    Table1Offset = 1
    Table2Offset = 1
    Table1Title = 1
    Table2Title = 1
    
    #Based on the table parameters we calculate corner positions of the tables
    #DO NOT change these ones!
    Table1LeftCorner = 1+Table1Offset
    #Where 1 is for the first cell index
    Table2LeftCorner = Table1LeftCorner+Table1Title+1+nofmethods+Table2Offset    
    #Where 1 is for the method enumeration line 
    
    ####################################################
    #Titles
    ####################################################
    #Table 1 Title
    sheet.merge_cells('A{}:{}{}'.format(Table1LeftCorner,format(chr(nofmethods + ord('A'))),Table1LeftCorner))
    sheet['A{}'.format(Table1LeftCorner)].style = Style(font=Font(bold=True))
    sheet['A{}'.format(Table1LeftCorner)] = 'Absolute Errors'    
    #Table 2 Title
    sheet.merge_cells('A{}:{}{}'.format(Table2LeftCorner,format(chr(nofmethods + ord('A'))),Table2LeftCorner))
    sheet['A{}'.format(Table2LeftCorner)].style = Style(font=Font(bold=True))
    sheet['A{}'.format(Table2LeftCorner)] = 'Relative Errors'
    #####################################################
    
    for i in range(len(ListOfMethods)):
        iPlusBchar = chr(i + ord('B')) 
        #Table 1
        sheet['A{}'.format((i+1)+Table1LeftCorner+Table1Title)] = ListOfMethods[i]
        sheet['{}{}'.format(iPlusBchar,Table1LeftCorner+Table1Title)] = ListOfMethods[i]
        sheet['{}{}'.format(iPlusBchar,(i+1)+Table1LeftCorner+Table1Title)] = 0
        #Table 2
        sheet['A{}'.format((i+1)+Table2LeftCorner+Table2Title)] = ListOfMethods[i]
        sheet['{}{}'.format(iPlusBchar,Table2LeftCorner+Table2Title)] = ListOfMethods[i]
        sheet['{}{}'.format(iPlusBchar,(i+1)+Table2LeftCorner+Table2Title)] = 0
    ####################################################
        
    for subset in itertools.combinations(enumerate(ListOfMethods),2):
        value1 = ListOfValues[subset[0][0]]
        value2 = ListOfValues[subset[1][0]]
        normofadiff = norm(value1-value2)
        normofrdiff1 = norm((value1-value2)/value1)
        normofrdiff2 = norm((value1-value2)/value2)
        
        if np.isinf(normofrdiff1):
            normofrdiff1 = 'inf'
        if np.isinf(normofrdiff2):
            normofrdiff2 = 'inf'
        
        print '\nAbsolute error for {} - {} pair is'.format(subset[0][1],subset[1][1])
        print normofadiff
        print 'Relative error for {} - {} pair is'.format(subset[0][1],subset[1][1])
        print normofrdiff1
        print 'Relative error for {} - {} pair is'.format(subset[1][1],subset[0][1])
        print normofrdiff2
        
        ###################################################
        # Save data to Table 1
        # Step 2a. Data Writing
        ###################################################
        idx1 = subset[0][0] 
        idx2 = subset[1][0]
        sheet['{}{}'.format(format(chr(idx1 + ord('B'))),(idx2+1)+Table1LeftCorner+Table1Title)] = normofadiff
        sheet['{}{}'.format(format(chr(idx2 + ord('B'))),(idx1+1)+Table1LeftCorner+Table1Title)] = normofadiff
        
        ###################################################
        # Save data to Table 2
        # Step 2b. Data Writing
        ################################################### 
        sheet['{}{}'.format(format(chr(idx1 + ord('B'))),(idx2+1)+Table2LeftCorner+Table2Title)] = normofrdiff1
        sheet['{}{}'.format(format(chr(idx2 + ord('B'))),(idx1+1)+Table2LeftCorner+Table2Title)] = normofrdiff2
        
    ###################################################
    # Write data to an excel file
    # Step 3. Saving data to the file
    #######################################################
    wb.save(fname)

def Speed_InitFirstSheet(wb,FirstSheetName,ListOfMethods):
    sheet = wb.get_sheet_by_name(FirstSheetName)
    #Main Table
    sheet['A1'].style = Style(font=Font(bold=True))
    sheet['A1'] = 'N'
    for i in range(len(ListOfMethods)):
        i4PlusBchar = GetString('B',i)
        sheet['{}1'.format(i4PlusBchar)].style = Style(font=Font(bold=True))
        sheet['{}1'.format(i4PlusBchar)] = ListOfMethods[i]
        
def Speed_SaveResults(wb, FirstSheetName, ListOfMethods, N, ListOfTimes):
    sheet = wb.get_sheet_by_name(FirstSheetName)
    #Main Table
    sheet['A{}'.format(N)] = N
    for i in range(len(ListOfMethods)):
        i4PlusBchar = GetString('B',i)
        sheet['{}{}'.format(i4PlusBchar, N)] = ListOfTimes[i]

def SaveSpeedtoExcelFile(fname,N,ListOfMethods,ListOfTimes):
    
    ###################################################
    # Write data to an excel file
    # Step 1. Initialization
    ###################################################
    FirstSheetName = fname.split('.')[0]
    
    if(os.path.isfile(fname)):     
        wb = openpyxl.load_workbook(fname)
        sheet = wb.get_sheet_by_name(FirstSheetName)
    else:
        wb = openpyxl.Workbook()
        #First sheet is the main one for the results
        sheet = wb.active
        sheet.title=FirstSheetName
        Speed_InitFirstSheet(wb,FirstSheetName,ListOfMethods)
    
    Speed_SaveResults(wb, FirstSheetName, ListOfMethods, N, ListOfTimes)    
    
    ###################################################
    # Write data to an excel file
    # Step 3. Saving data to the file
    #######################################################
    wb.save(fname)    